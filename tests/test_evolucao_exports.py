from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

import app1


def _find_status_row(ws, instituicao: str, periodo: str, indicador: str) -> int:
    for row_idx in range(2, ws.max_row + 1):
        if (
            ws.cell(row=row_idx, column=1).value == instituicao
            and ws.cell(row=row_idx, column=2).value == periodo
            and ws.cell(row=row_idx, column=3).value == indicador
        ):
            return row_idx
    raise AssertionError(f"status row not found: {instituicao=} {periodo=} {indicador=}")


def test_evolucao_visual_export_includes_analytical_status_sheet():
    df_show = pd.DataFrame(
        [
            {"Métrica": "ROE Ac. Anualizado (%)", "Dez/25": "10,00%"},
            {"Métrica": "Índice de Capital Principal (CET1)", "Dez/25": "-"},
        ]
    )
    df_ano = pd.DataFrame(
        [
            {
                "Instituição": "TEST BANK - PRUDENCIAL",
                "Período": "4/2025",
                "Lucro Líquido Acumulado YTD": 10.0,
                "Patrimônio Líquido": 100.0,
                "Carteira de Crédito Bruta": 500.0,
                "Core Funding": 400.0,
                "ROE Ac. Anualizado (%)": 0.10,
                "Carteira de Crédito Bruta / PL": 5.0,
                "Índice de Capital Principal (CET1)": None,
                "Índice de Capital T1 (%)": None,
                "Índice de Basileia Total (%)": None,
            }
        ]
    )
    core_funding_trace_map = {
        "4/2025": {
            "status": "fallback_capitacoes_view",
            "source": "Captações (base da aba Evolução)",
            "note": "A visualização caiu defensivamente para Captações porque a reconstrução completa pelo Rel. 3 não ficou disponível.",
        }
    }
    carteira_trace_map = {
        "4/2025": {
            "status": "fallback_net_components",
            "source": "Operações de Crédito (e) + Operações de Arrendamento Financeiro (f) + Outras Operações com Características de Concessão de Crédito (g) + Valores a Receber de Transações de Pagamentos - Usuários Finais (Pós-pago) (h)",
            "note": "Sem componentes brutos completos para a regra canônica do período; o valor caiu explicitamente para a soma líquida e+f+g+h.",
        }
    }

    output = app1._gerar_excel_evolucao_tabela_visual(
        df_show=df_show,
        periodos_cols=["Dez/25"],
        instituicao="TEST BANK - PRUDENCIAL",
        periodo_inicio="4/2025",
        periodo_final="4/2025",
        df_ano=df_ano,
        core_funding_trace_map=core_funding_trace_map,
        carteira_trace_map=carteira_trace_map,
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["status_analitico"]

    core_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "Core Funding*")
    cart_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "Carteira de Crédito*")
    cet1_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "Índice de Capital Principal (CET1)")

    assert ws.cell(row=core_row, column=5).value == "fallback_capitacoes_view"
    assert "captações" in str(ws.cell(row=core_row, column=7).value).lower()
    assert ws.cell(row=cart_row, column=5).value == "fallback_net_components"
    assert "e+f+g+h" in str(ws.cell(row=cart_row, column=7).value).lower()

    assert ws.cell(row=cet1_row, column=5).value == "capital_report_missing"
    assert "rel. 5" in str(ws.cell(row=cet1_row, column=7).value).lower()


def test_decorate_evolucao_visual_table_marks_fallback_and_unavailability():
    df_show = pd.DataFrame(
        [
            {"Métrica": "Carteira de Crédito*", "Dez/25": "500"},
            {"Métrica": "Core Funding*", "Dez/25": "400"},
            {"Métrica": "Índice de Capital Principal (CET1)", "Dez/25": "-"},
        ]
    )
    df_ano = pd.DataFrame(
        [
            {
                "Instituição": "TEST BANK - PRUDENCIAL",
                "Período": "4/2025",
                "Carteira de Crédito Bruta": 500.0,
                "Core Funding": 400.0,
                "Índice de Capital Principal (CET1)": None,
            }
        ]
    )
    core_funding_trace_map = {
        "4/2025": {
            "status": "fallback_capitacoes_view",
            "source": "Captações (base da aba Evolução)",
            "note": "A visualização caiu defensivamente para Captações porque a reconstrução completa pelo Rel. 3 não ficou disponível.",
        }
    }
    carteira_trace_map = {
        "4/2025": {
            "status": "fallback_net_components",
            "source": "Fallback líquido e+f+g+h",
            "note": "Sem componentes brutos completos para a regra canônica do período; o valor caiu explicitamente para a soma líquida e+f+g+h.",
        }
    }

    decorated, tooltips, marker_presence = app1._decorate_evolucao_visual_table(
        df_show,
        ["Dez/25"],
        instituicao="TEST BANK - PRUDENCIAL",
        df_ano=df_ano,
        core_funding_trace_map=core_funding_trace_map,
        carteira_trace_map=carteira_trace_map,
    )

    assert decorated.loc[0, "Dez/25"] == "500*"
    assert decorated.loc[1, "Dez/25"] == "400*"
    assert decorated.loc[2, "Dez/25"] == "-†"
    assert marker_presence["fallback"] is True
    assert marker_presence["unavailable"] is True
    assert "fallback_net_components" in tooltips[("Carteira de Crédito*", "Dez/25")]
    assert "fallback_capitacoes_view" in tooltips[("Core Funding*", "Dez/25")]
    assert "capital_report_missing" in tooltips[("Índice de Capital Principal (CET1)", "Dez/25")]


def test_evolucao_visual_export_appends_marker_legend():
    df_show = pd.DataFrame(
        [
            {"Métrica": "Core Funding*", "Dez/25": "400*"},
            {"Métrica": "Índice de Capital Principal (CET1)", "Dez/25": "-†"},
        ]
    )

    output = app1._gerar_excel_evolucao_tabela_visual(
        df_show=df_show,
        periodos_cols=["Dez/25"],
        instituicao="TEST BANK - PRUDENCIAL",
        periodo_inicio="4/2025",
        periodo_final="4/2025",
        marker_presence={"fallback": True, "unavailable": True},
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["evolucao_visual"]

    assert "símbolos visuais" in str(ws.cell(row=2, column=1).value).lower()
    assert "status_analitico" in str(ws.cell(row=2, column=1).value).lower()
    assert ws.cell(row=4, column=2).value == "400*"
    assert ws.cell(row=5, column=2).value == "-†"
