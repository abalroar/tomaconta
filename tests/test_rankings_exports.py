from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

import app1


def _find_status_row(ws, instituicao: str, periodo: str, indicador: str) -> int:
    for row_idx in range(2, ws.max_row + 1):
        if (
            ws.cell(row=row_idx, column=1).value == instituicao
            and ws.cell(row=row_idx, column=2).value == periodo
            and ws.cell(row=row_idx, column=3).value == indicador
        ):
            return row_idx
    raise AssertionError(f"status row not found: {instituicao=} {periodo=} {indicador=}")


def test_rankings_raw_export_includes_included_and_omitted_institutions():
    df_export = pd.DataFrame(
        [
            {
                "Período": "4/2025",
                "Instituição": "BANK A - PRUDENCIAL",
                "Indicador": "Ativo Total",
                "Valor": 1000.0,
                "Ranking": 1,
                "Média do Grupo": 1000.0,
                "Tipo de Média": "Média simples",
                "Diferença vs Média": 0.0,
            }
        ]
    )
    df_source = pd.DataFrame(
        [
            {"Instituição": "BANK A - PRUDENCIAL", "Período": "4/2025", "Ativo Total": 1000.0},
            {"Instituição": "BANK B - PRUDENCIAL", "Período": "4/2025", "Ativo Total": None},
        ]
    )

    output = app1._gerar_excel_rankings_dados_puros(
        df_export=df_export,
        periodo_ref="4/2025",
        indicador_label="Ativo Total",
        indicador_col="Ativo Total",
        bancos_referencia=[
            "BANK A - PRUDENCIAL",
            "BANK B - PRUDENCIAL",
            "BANK C - PRUDENCIAL",
        ],
        df_source=df_source,
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["status_analitico"]

    included_row = _find_status_row(ws, "BANK A - PRUDENCIAL", "Dez/25", "Ativo Total")
    missing_row = _find_status_row(ws, "BANK B - PRUDENCIAL", "Dez/25", "Ativo Total")
    absent_row = _find_status_row(ws, "BANK C - PRUDENCIAL", "Dez/25", "Ativo Total")

    assert ws.cell(row=included_row, column=5).value == "included"
    assert ws.cell(row=included_row, column=6).value == "Ativo Total"

    assert ws.cell(row=missing_row, column=5).value == "indicator_missing"
    assert "omitida" in str(ws.cell(row=missing_row, column=7).value).lower()

    assert ws.cell(row=absent_row, column=5).value == "institution_period_missing"
    assert "dataset filtrado" in str(ws.cell(row=absent_row, column=7).value).lower()


def test_rankings_capital_raw_export_includes_missing_rel5_status():
    df_export = pd.DataFrame(
        [
            {
                "Período": "4/2025",
                "Instituição": "BANK A - PRUDENCIAL",
                "CET1 (%)": 11.2,
                "AT1 (%)": 1.0,
                "T2 (%)": 2.0,
                "Índice de Basileia Total (%)": 14.2,
                "Ranking": 1,
                "Diferença vs Média (%)": 0.0,
            }
        ]
    )
    df_source = pd.DataFrame(
        [
            {"Instituição": "BANK A - PRUDENCIAL", "Período": "4/2025", "Índice de Basileia Total (%)": 14.2},
            {"Instituição": "BANK B - PRUDENCIAL", "Período": "4/2025", "Índice de Basileia Total (%)": None},
        ]
    )

    output = app1._gerar_excel_rankings_capital_dados_puros(
        df_export=df_export,
        periodos_ref=["4/2025"],
        bancos_referencia=[
            "BANK A - PRUDENCIAL",
            "BANK B - PRUDENCIAL",
            "BANK C - PRUDENCIAL",
        ],
        df_source=df_source,
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["status_analitico"]

    included_row = _find_status_row(ws, "BANK A - PRUDENCIAL", "Dez/25", "Índice de Basileia Total (%)")
    missing_row = _find_status_row(ws, "BANK B - PRUDENCIAL", "Dez/25", "Índice de Basileia Total (%)")
    absent_row = _find_status_row(ws, "BANK C - PRUDENCIAL", "Dez/25", "Índice de Basileia Total (%)")

    assert ws.cell(row=included_row, column=5).value == "included"
    assert "rel. 5" in str(ws.cell(row=included_row, column=6).value).lower()

    assert ws.cell(row=missing_row, column=5).value == "capital_indicator_missing"
    assert "basileia" in str(ws.cell(row=missing_row, column=7).value).lower()

    assert ws.cell(row=absent_row, column=5).value == "institution_period_missing"
    assert "rel. 5" in str(ws.cell(row=absent_row, column=7).value).lower()


def test_rankings_export_custo_credito_usa_percentual_e_celula_vazia():
    """Custo de Crédito exporta valor bruto decimal com formato %, e N/D como célula vazia."""
    indicador = app1.METRIC_CUSTO_CREDITO
    df_export = pd.DataFrame(
        [
            {
                "Período": "1/2026",
                "Instituição": "BANK A - PRUDENCIAL",
                "Indicador": indicador,
                "Valor": 0.0306,
                "Ranking": 1,
            }
        ]
    )
    df_source = pd.DataFrame(
        [
            {"Instituição": "BANK A - PRUDENCIAL", "Período": "1/2026", indicador: 0.0306},
            {"Instituição": "BANK B - PRUDENCIAL", "Período": "1/2026", indicador: None},
        ]
    )

    output = app1._gerar_excel_rankings_dados_puros(
        df_export=df_export,
        periodo_ref="1/2026",
        indicador_label=indicador,
        indicador_col=indicador,
        bancos_referencia=["BANK A - PRUDENCIAL", "BANK B - PRUDENCIAL"],
        df_source=df_source,
    )

    ws = load_workbook(BytesIO(output.getvalue()), data_only=True)["status_analitico"]
    incluida = _find_status_row(ws, "BANK A - PRUDENCIAL", "Mar/26", indicador)
    omitida = _find_status_row(ws, "BANK B - PRUDENCIAL", "Mar/26", indicador)

    celula_incluida = ws.cell(row=incluida, column=4)
    assert celula_incluida.value == pytest.approx(0.0306)
    assert "%" in str(celula_incluida.number_format)

    assert ws.cell(row=omitida, column=5).value == "indicator_missing"
    assert ws.cell(row=omitida, column=4).value is None
