from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

import app1
from utils.ifdata_cache import CacheManager


def _find_row_by_label(ws, label: str) -> int:
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == label:
            return row_idx
    raise AssertionError(f"label not found: {label}")


def _find_status_row(ws, instituicao: str, periodo: str, indicador: str) -> int:
    for row_idx in range(2, ws.max_row + 1):
        if (
            ws.cell(row=row_idx, column=1).value == instituicao
            and ws.cell(row=row_idx, column=2).value == periodo
            and ws.cell(row=row_idx, column=3).value == indicador
        ):
            return row_idx
    raise AssertionError(f"status row not found: {instituicao=} {periodo=} {indicador=}")


def test_peers_raw_export_writes_numeric_values_without_visual_arrows():
    bancos = ["TEST BANK - PRUDENCIAL"]
    periodos = ["4/2025"]
    valores = {
        ("Ativo Total", "TEST BANK - PRUDENCIAL", "4/2025"): 1_500_000_000.0,
        ("ROE Acumulado YTD (%)", "TEST BANK - PRUDENCIAL", "4/2025"): 0.125,
    }
    colunas_usadas = {
        "Ativo Total": "Ativo Total",
        "ROE Acumulado YTD (%)": "ROE Ac. Anualizado (%)",
    }
    delta_flags = {
        ("Ativo Total", "TEST BANK - PRUDENCIAL", "4/2025"): "up",
        ("ROE Acumulado YTD (%)", "TEST BANK - PRUDENCIAL", "4/2025"): "down",
    }

    output = app1._gerar_excel_peers_dados_puros(
        bancos=bancos,
        periodos=periodos,
        valores=valores,
        colunas_usadas=colunas_usadas,
        delta_flags=delta_flags,
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["dados_numericos"]

    ativo_row = _find_row_by_label(ws, "Ativo Total")
    roe_row = _find_row_by_label(ws, "ROE Acumulado YTD (%)")

    ativo_val = ws.cell(row=ativo_row, column=2).value
    roe_val = ws.cell(row=roe_row, column=2).value

    assert isinstance(ativo_val, (int, float))
    assert isinstance(roe_val, (int, float))
    assert ativo_val == 1_500_000_000.0
    assert round(float(roe_val), 6) == 0.125


def test_peers_raw_export_includes_analytical_status_sheet():
    bancos = ["TEST BANK - PRUDENCIAL"]
    periodos = ["4/2025"]
    valores = {
        ("Depósitos Totais", "TEST BANK - PRUDENCIAL", "4/2025"): 106.0,
        ("Core Funding*", "TEST BANK - PRUDENCIAL", "4/2025"): None,
    }
    colunas_usadas = {
        "Depósitos Totais": "Depósitos Totais",
        "Core Funding*": "Core Funding*",
    }
    delta_flags = {}
    df_base = pd.DataFrame(
        [
            {
                "Instituição": "TEST BANK - PRUDENCIAL",
                "Período": "4/2025",
                "Depósitos Totais": 106.0,
                "Trace::Depósitos Totais::Status": "fallback_components",
                "Trace::Depósitos Totais::Campo Selecionado": None,
                "Core Funding*": None,
                "Trace::Core Funding::Status": "missing_required_component",
                "Trace::Core Funding::Campo Selecionado": "Captações (e) + Instrumentos de Dívida Elegíveis a Capital (h)",
                "Trace::Core Funding::Captações (e)": 210.0,
                "Trace::Core Funding::Instrumentos de Dívida Elegíveis a Capital (h)": None,
            }
        ]
    )

    output = app1._gerar_excel_peers_dados_puros(
        bancos=bancos,
        periodos=periodos,
        valores=valores,
        colunas_usadas=colunas_usadas,
        delta_flags=delta_flags,
        df_base=df_base,
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["status_analitico"]

    dep_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "Depósitos Totais")
    core_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "Core Funding*")

    assert ws.cell(row=dep_row, column=5).value == "fallback_components"
    assert ws.cell(row=dep_row, column=6).value == "Soma dos subtipos a1..a6"
    assert "soma dos subtipos" in str(ws.cell(row=dep_row, column=7).value).lower()

    assert ws.cell(row=core_row, column=5).value == "missing_required_component"
    assert ws.cell(row=core_row, column=6).value == "Captações (e) + Instrumentos de Dívida Elegíveis a Capital (h)"
    assert "indisponível" in str(ws.cell(row=core_row, column=7).value).lower()


def test_evolucao_raw_export_includes_analytical_status_sheet():
    df_graph = pd.DataFrame(
        [
            {
                "Ano": 2025,
                "Lucro Líquido": 10.0,
                "Patrimônio Líquido": 100.0,
                "Carteira de Crédito*": 500.0,
                "Core Funding*": 400.0,
            }
        ]
    )
    df_metric = pd.DataFrame(
        [
            {
                "Métrica": "ROE Ac. Anualizado (%)",
                "Dez/25": 0.10,
            },
            {
                "Métrica": "Índice de Capital Principal (CET1)",
                "Dez/25": None,
            },
        ]
    )
    df_ano = pd.DataFrame(
        [
            {
                "Instituição": "TEST BANK - PRUDENCIAL",
                "Período": "4/2025",
                "Lucro Líquido Acumulado YTD": 10.0,
                "Patrimônio Líquido": 100.0,
                "Carteira de Crédito Bruta": 500.0,
                "Core Funding": 400.0,
                "ROE Ac. Anualizado (%)": 0.10,
                "Carteira de Crédito Bruta / PL": 5.0,
                "Índice de Capital Principal (CET1)": None,
                "Índice de Capital T1 (%)": None,
                "Índice de Basileia Total (%)": None,
            }
        ]
    )
    core_funding_trace_map = {
        "4/2025": {
            "status": "fallback_capitacoes_view",
            "source": "Captações (base da aba Evolução)",
            "note": "A visualização caiu defensivamente para Captações porque a reconstrução completa pelo Rel. 3 não ficou disponível.",
        }
    }

    output = app1._gerar_excel_evolucao_dados_puros(
        instituicao="TEST BANK - PRUDENCIAL",
        df_graph=df_graph,
        df_metric=df_metric,
        df_ano=df_ano,
        core_funding_trace_map=core_funding_trace_map,
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["status_analitico"]

    core_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "Core Funding*")
    cet1_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "Índice de Capital Principal (CET1)")
    roe_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "ROE Ac. Anualizado (%)")

    assert ws.cell(row=core_row, column=5).value == "fallback_capitacoes_view"
    assert ws.cell(row=core_row, column=6).value == "Captações (base da aba Evolução)"
    assert "captações" in str(ws.cell(row=core_row, column=7).value).lower()

    assert ws.cell(row=cet1_row, column=5).value == "capital_report_missing"
    assert "rel. 5" in str(ws.cell(row=cet1_row, column=7).value).lower()

    assert ws.cell(row=roe_row, column=5).value == "derived_from_curated"
    assert "pl médio" in str(ws.cell(row=roe_row, column=6).value).lower()


def test_peers_ui_tooltips_include_fallback_and_unavailability_markers():
    bancos = ["TEST BANK - PRUDENCIAL"]
    periodos = ["4/2025"]
    valores = {
        ("Depósitos Totais", "TEST BANK - PRUDENCIAL", "4/2025"): 106.0,
        ("Perda Esperada / Estágio 3", "TEST BANK - PRUDENCIAL", "4/2025"): None,
    }
    colunas_usadas = {
        "Depósitos Totais": "Depósitos Totais",
        "Perda Esperada / Estágio 3": "Perda Esperada / Estágio 3",
    }
    df_base = pd.DataFrame(
        [
            {
                "Instituição": "TEST BANK - PRUDENCIAL",
                "Período": "4/2025",
                "Depósitos Totais": 106.0,
                "Trace::Depósitos Totais::Status": "fallback_components",
                "Trace::Depósitos Totais::Campo Selecionado": None,
                "Perda Esperada / Estágio 3": None,
                "Trace::Qualidade Carteira::Status": "source_structurally_unavailable",
            }
        ]
    )

    status_lookup = app1._build_peers_status_lookup(
        df_base=df_base,
        bancos=bancos,
        periodos=periodos,
        valores=valores,
        colunas_usadas=colunas_usadas,
    )
    tooltips_ui, status_markers, marker_presence = app1._merge_peers_analytical_tooltips(
        tooltips={},
        status_lookup=status_lookup,
    )
    html = app1._render_peers_table_html(
        bancos,
        periodos,
        valores,
        colunas_usadas,
        delta_flags={},
        delta_context={},
        tooltips=tooltips_ui,
        status_markers=status_markers,
    )

    assert status_markers[("Depósitos Totais", "TEST BANK - PRUDENCIAL", "4/2025")] == "*"
    assert status_markers[("Perda Esperada / Estágio 3", "TEST BANK - PRUDENCIAL", "4/2025")] == "†"
    assert marker_presence["fallback"] is True
    assert marker_presence["unavailable"] is True
    assert "fallback_components" in tooltips_ui[("Depósitos Totais", "TEST BANK - PRUDENCIAL", "4/2025")]
    assert "source_structurally_unavailable" in tooltips_ui[("Perda Esperada / Estágio 3", "TEST BANK - PRUDENCIAL", "4/2025")]
    assert "N/D" in html
    assert "status-mark--fallback" in html
    assert "status-mark--unavailable" in html


def test_peers_visual_export_also_includes_analytical_status_sheet():
    bancos = ["TEST BANK - PRUDENCIAL"]
    periodos = ["4/2025"]
    valores = {
        ("Depósitos Totais", "TEST BANK - PRUDENCIAL", "4/2025"): 106.0,
    }
    colunas_usadas = {"Depósitos Totais": "Depósitos Totais"}
    df_base = pd.DataFrame(
        [
            {
                "Instituição": "TEST BANK - PRUDENCIAL",
                "Período": "4/2025",
                "Depósitos Totais": 106.0,
                "Trace::Depósitos Totais::Status": "fallback_components",
                "Trace::Depósitos Totais::Campo Selecionado": None,
            }
        ]
    )

    output = app1._gerar_excel_peers_tabela(
        bancos=bancos,
        periodos=periodos,
        valores=valores,
        colunas_usadas=colunas_usadas,
        delta_flags={},
        df_base=df_base,
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["status_analitico"]
    dep_row = _find_status_row(ws, "TEST BANK - PRUDENCIAL", "Dez/25", "Depósitos Totais")

    assert ws.cell(row=dep_row, column=5).value == "fallback_components"
    assert "subtipos" in str(ws.cell(row=dep_row, column=7).value).lower()


def test_peers_visual_export_marks_cells_and_note_with_analytical_symbols():
    bancos = ["TEST BANK - PRUDENCIAL"]
    periodos = ["4/2025"]
    valores = {
        ("Depósitos Totais", "TEST BANK - PRUDENCIAL", "4/2025"): 106.0,
        ("Perda Esperada / Estágio 3", "TEST BANK - PRUDENCIAL", "4/2025"): None,
    }
    colunas_usadas = {
        "Depósitos Totais": "Depósitos Totais",
        "Perda Esperada / Estágio 3": "Perda Esperada / Estágio 3",
    }
    df_base = pd.DataFrame(
        [
            {
                "Instituição": "TEST BANK - PRUDENCIAL",
                "Período": "4/2025",
                "Depósitos Totais": 106.0,
                "Trace::Depósitos Totais::Status": "fallback_components",
                "Trace::Depósitos Totais::Campo Selecionado": None,
                "Perda Esperada / Estágio 3": None,
                "Trace::Qualidade Carteira::Status": "source_structurally_unavailable",
            }
        ]
    )

    output = app1._gerar_excel_peers_tabela(
        bancos=bancos,
        periodos=periodos,
        valores=valores,
        colunas_usadas=colunas_usadas,
        delta_flags={},
        df_base=df_base,
    )

    wb = load_workbook(BytesIO(output.getvalue()), data_only=True)
    ws = wb["peers_tabela"]

    dep_row = _find_row_by_label(ws, "Depósitos Totais")
    est3_row = _find_row_by_label(ws, "Perda Esperada / Estágio 3")

    assert ws.cell(row=3, column=1).value is not None
    assert "símbolos visuais" in str(ws.cell(row=3, column=1).value).lower()
    assert "status_analitico" in str(ws.cell(row=3, column=1).value).lower()
    assert str(ws.cell(row=dep_row, column=2).value).endswith("*")
    assert str(ws.cell(row=est3_row, column=2).value).endswith("†")


def test_decorate_peers_visual_value_preserves_marker_and_delta_order():
    assert app1._decorate_peers_visual_value("123,45 MM", status_marker="*", delta_flag="up") == "123,45 MM* ▲"
    assert app1._decorate_peers_visual_value("N/D", status_marker="†", delta_flag=None) == "N/D†"


def test_peers_base_dre_options_always_include_individual():
    assert app1.PEERS_BASE_DRE_OPTIONS == ["Consolidada / Prudencial", "Individual"]
    assert app1._normalizar_base_dre_peers("Consolidada") == "Consolidada / Prudencial"
    assert app1._normalizar_base_dre_peers("Individual") == "Individual"


def test_individual_caches_use_official_ifdata_type_three():
    manager = CacheManager()

    assert manager.get_cache("principal_individual").tipo_instituicao == 3
    assert manager.get_cache("dre_individual").tipo_instituicao == 3


def test_peers_individual_identity_keeps_historical_rows_under_latest_name():
    df = pd.DataFrame(
        [
            {"CodInst": "C001", "Instituição": "Banco Antigo", "Período": "1/2025", "Ativo Total": 90.0},
            {"CodInst": "C001", "Instituição": "Banco Atual", "Período": "1/2026", "Ativo Total": 100.0},
        ]
    )

    out = app1._apply_peers_individual_display_names(df, {"C001": "Banco Atual"})

    assert out["Instituição"].tolist() == ["Banco Atual", "Banco Atual"]
    assert app1._obter_valor_peers(out, "Banco Atual", "1/2025", "Ativo Total") == 90.0
    assert app1._obter_valor_peers(out, "Banco Atual", "1/2026", "Ativo Total") == 100.0


def test_peers_individual_uses_classified_credit_before_cosif_2025():
    df = pd.DataFrame(
        [
            {
                "CodInst": "001",
                "Instituição": "Banco A",
                "Período": "4/2024",
                "Carteira de Crédito Classificada": 123.0,
                "Carteira de Crédito": None,
            }
        ]
    )

    extra = app1._preparar_metricas_extra_peers_individual_from_slice(
        df,
        ["Banco A"],
        ["4/2024"],
    )

    assert extra["Carteira de Crédito*"][("Banco A", "4/2024")] == 123.0


def test_peers_slice_recomputes_loss_coverage_and_formats_above_100_percent():
    df_slice = pd.DataFrame(
        [
            {
                "Instituição": "GM - PRUDENCIAL",
                "Período": "1/2026",
                "Perda Esperada": -399.13,
                "Ativos Estágio 3": 333.2,
                "Perda Esperada / Estágio 3": -1.197869,
                "Carteira de Crédito Bruta": 1000.0,
                "Inadimplência 4.966": 20.0,
            }
        ]
    )

    extra = app1._preparar_metricas_extra_peers_from_slice(
        df_slice,
        ["GM - PRUDENCIAL"],
        ["1/2026"],
    )

    chave = ("GM - PRUDENCIAL", "1/2026")
    assert round(extra["Perda Esperada / Estágio 3"][chave], 6) == round(399.13 / 333.2, 6)
    assert round(extra["Inadimplência / Carteira de Crédito"][chave], 6) == round(20.0 / 1000.0, 6)
    assert app1._formatar_valor_peers(extra["Perda Esperada / Estágio 3"][chave], "Perda Esperada / Estágio 3") == "119,79%"


def test_peers_generic_missing_value_gets_unavailability_note_and_marker():
    bancos = ["TEST BANK - PRUDENCIAL"]
    periodos = ["4/2025"]
    valores = {
        ("Ativo Total", "TEST BANK - PRUDENCIAL", "4/2025"): None,
    }
    colunas_usadas = {"Ativo Total": "Ativo Total"}
    df_base = pd.DataFrame(
        [
            {
                "Instituição": "TEST BANK - PRUDENCIAL",
                "Período": "4/2025",
                "Ativo Total": None,
            }
        ]
    )

    status_lookup = app1._build_peers_status_lookup(
        df_base=df_base,
        bancos=bancos,
        periodos=periodos,
        valores=valores,
        colunas_usadas=colunas_usadas,
    )
    tooltips_ui, status_markers, marker_presence = app1._merge_peers_analytical_tooltips(
        tooltips={},
        status_lookup=status_lookup,
    )

    assert status_lookup[("Ativo Total", "TEST BANK - PRUDENCIAL", "4/2025")]["Status analítico"] == "missing"
    assert "base curada" in str(status_lookup[("Ativo Total", "TEST BANK - PRUDENCIAL", "4/2025")]["Observação"]).lower()
    assert status_markers[("Ativo Total", "TEST BANK - PRUDENCIAL", "4/2025")] == "†"
    assert marker_presence["unavailable"] is True
    assert "sem valor disponível" in tooltips_ui[("Ativo Total", "TEST BANK - PRUDENCIAL", "4/2025")].lower()
