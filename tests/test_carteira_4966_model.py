import html
import re
from io import BytesIO

import openpyxl
import pandas as pd
import pytest

from tabs.carteira_4966 import (
    EXPECTED_LOSS_COLUMNS,
    QualityIssue,
    ROW_SPECS,
    TITLE,
    build_carteira_4966_excel,
    build_carteira_4966_model,
    build_carteira_4966_raw_excel,
    format_brl_millions,
    format_percentage,
    model_to_audit_dataframe,
    quality_issue_message,
    quality_issues_dataframe,
    render_carteira_4966_html,
)


MM = 1_000_000.0


def _carteira_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Instituição": ["BANCO TESTE - PRUDENCIAL"] * 4,
            "Período": ["1/2025", "2/2025", "4/2025", "1/2026"],
            "C1": [10, 12, 20, 22],
            "C2": [10, 12, 20, 22],
            "C3": [10, 12, 20, 22],
            "C4": [20, 24, 40, 44],
            "C5": [50, 60, 100, 110],
            "Total não Individualizado": [5, 6, 10, 11],
            "Carteira não Informada ou não se Aplica": [10, 12, 20, 22],
            "Total Exterior": [20, 24, 40, 44],
            "Total Geral": [135, 162, 270, 297],
            "Inadimplência": [10, None, 30, 33],
        }
    ).assign(
        **{
            column: lambda frame, column=column: frame[column] * MM
            for column in [
                "C1",
                "C2",
                "C3",
                "C4",
                "C5",
                "Total não Individualizado",
                "Carteira não Informada ou não se Aplica",
                "Total Exterior",
                "Total Geral",
                "Inadimplência",
            ]
        }
    )


def _ativo_frame() -> pd.DataFrame:
    rows = []
    for period, multiplier in zip(["1/2025", "2/2025", "4/2025", "1/2026"], [1, 1.2, 2, 2.2]):
        row = {
            "Instituição": "BANCO TESTE - PRUDENCIAL",
            "Período": period,
            EXPECTED_LOSS_COLUMNS[0]: -10 * multiplier * MM,
            EXPECTED_LOSS_COLUMNS[1]: -2 * multiplier * MM,
            EXPECTED_LOSS_COLUMNS[2]: -3 * multiplier * MM,
            EXPECTED_LOSS_COLUMNS[3]: -5 * multiplier * MM,
            "Hedge de Valor Justo (e3)": 100 * multiplier * MM,
            "Ajuste a Valor Justo (g4)": 50 * multiplier * MM,
        }
        rows.append(row)
    return pd.DataFrame(rows)


def _model():
    return build_carteira_4966_model(
        _carteira_frame(),
        _ativo_frame(),
        ["1/2026", "1/2025", "4/2025", "2/2025"],
    )


def _html_row(rendered: str, label: str) -> str:
    match = re.search(
        rf"<tr[^>]*>.*?<th[^>]*>{re.escape(label)}</th>(.*?)</tr>",
        rendered,
        flags=re.DOTALL,
    )
    assert match is not None, f"Linha {label!r} não encontrada no HTML"
    return match.group(1)


def _visible_text(fragment: str) -> str:
    return re.sub(r"<[^>]+>", "", html.unescape(fragment)).replace("\n", "").strip()


def _excel_value_cell(sheet, label: str):
    labels = {
        sheet.cell(row=row, column=1).value: row
        for row in range(1, sheet.max_row + 1)
    }
    assert label in labels, f"Linha {label!r} não encontrada no Excel"
    return sheet.cell(row=labels[label], column=2)


def test_model_uses_latest_q4_as_common_base_and_period_specific_ratios():
    model = _model()

    assert model.periods == ("1/2025", "2/2025", "4/2025", "1/2026")
    assert model.base_period == "4/2025"
    assert model.base_value == pytest.approx(270 * MM)
    assert model.cells["total_portfolio"]["1/2025"].secondary == pytest.approx(135 / 270)
    assert model.cells["c5"]["1/2025"].secondary == pytest.approx(50 / 270)
    assert model.cells["delinquency"]["1/2025"].secondary == pytest.approx(10 / 135)
    assert model.cells["delinquency"]["2/2025"].primary is None
    assert model.cells["delinquency"]["2/2025"].secondary is None
    assert model.qoq["2/2025"] == pytest.approx(162 / 135 - 1)
    assert model.qoq["1/2026"] == pytest.approx(297 / 270 - 1)


def test_provision_is_literal_expected_loss_and_excludes_hedge_and_fair_value():
    model = _model()

    provision = model.cells["provision"]["4/2025"].primary
    assert provision == pytest.approx(40 * MM)
    assert model.cells["provision_over_portfolio"]["4/2025"].primary == pytest.approx(40 / 270)
    assert model.cells["provision_over_c5"]["4/2025"].primary == pytest.approx(40 / 100)
    assert model.cells["provision_over_delinquency"]["4/2025"].primary == pytest.approx(40 / 30)
    assert model.cell_quality_issues("provision", "4/2025") == ()
    assert model.cell_quality_issues("provision_over_portfolio", "4/2025") == ()


def test_missing_source_and_zero_denominator_stay_distinct_from_published_zero():
    carteira = _carteira_frame().query("`Período` == '1/2025'").copy()
    carteira.loc[:, "C5"] = 0.0
    carteira.loc[:, "Inadimplência"] = 0.0
    carteira.loc[:, "Total Geral"] = 0.0
    ativo = _ativo_frame().query("`Período` == '1/2025'").copy()
    for column in EXPECTED_LOSS_COLUMNS:
        ativo.loc[:, column] = 0.0

    model = build_carteira_4966_model(carteira, ativo, ["1/2025"])

    assert model.cells["provision"]["1/2025"].primary == 0.0
    assert model.cells["provision_over_portfolio"]["1/2025"].primary is None
    assert model.cells["provision_over_c5"]["1/2025"].primary is None
    assert model.cells["provision_over_delinquency"]["1/2025"].primary is None

    without_ativo = build_carteira_4966_model(carteira, pd.DataFrame(), ["1/2025"])
    assert without_ativo.missing_provision_periods == ("1/2025",)


def test_without_q4_the_latest_displayed_period_becomes_the_base():
    model = build_carteira_4966_model(
        _carteira_frame(),
        _ativo_frame(),
        ["1/2025", "2/2025"],
    )

    assert model.base_period == "2/2025"
    assert model.cells["total_portfolio"]["2/2025"].secondary == pytest.approx(1.0)


def test_html_has_clean_grid_without_decorative_hatched_column():
    model = _model()
    rendered = render_carteira_4966_html(model)

    assert TITLE in rendered
    assert "repeating-linear-gradient" not in rendered
    assert "tc-4966-marker" not in rendered
    assert '<col style="width:14px">' not in rendered
    assert "border-collapse: collapse" in rendered
    assert "--tc-line: #dddddd" in rendered
    assert 'role="region"' in rendered
    assert 'tabindex="0"' in rendered
    assert rendered.count('<tbody data-group=') == 3
    assert 'colspan="2"' in rendered
    assert "Vencidos acima de 90 dias (conceito de arrasto)" in rendered
    assert "N/D" in rendered
    assert "PDD / Carteira Total (%)" in rendered
    assert "7,41%" in rendered
    assert "14,81%" in rendered
    assert "&gt;" not in rendered
    assert "—" not in rendered


def test_quality_tooltip_is_visible_on_hover_and_keyboard_focus():
    rendered = render_carteira_4966_html(_model_with_pdd_ratio(60.0))
    row = _html_row(rendered, "PDD / Carteira Total (%)")

    described_by = re.search(r'aria-describedby="([^"]+)"', row)
    assert described_by is not None
    assert f'id="{described_by.group(1)}"' in row
    assert 'role="tooltip"' in row
    assert 'tabindex="0"' in row
    assert 'data-quality="warning"' in row
    assert ' title=' not in row
    assert ".tc-4966-has-tooltip:hover .tc-4966-tooltip" in rendered
    assert ".tc-4966-has-tooltip:focus .tc-4966-tooltip" in rendered
    assert "Passe o cursor ou use a tecla Tab" in rendered

    reliable_row = _html_row(rendered, "C1")
    assert "tc-4966-quality-warning" not in reliable_row
    assert 'role="tooltip"' not in reliable_row


def test_paired_quality_issue_marks_both_subcells_in_html_and_visual_excel():
    model = _model()
    model.quality_issues = (
        QualityIssue(
            period="2/2025",
            severity="warning",
            code="missing_denominator",
            pdd_value=24 * MM,
            portfolio_value=None,
            ratio=None,
            row_key="delinquency",
            denominator_label="Inadimplência do Relatório 16",
        ),
    )

    row = _html_row(
        render_carteira_4966_html(model),
        "Vencidos acima de 90 dias (conceito de arrasto)",
    )
    assert row.count("tc-4966-quality-warning") == 2
    assert row.count('role="tooltip"') == 2
    assert row.count('tabindex="0"') == 2
    assert _visible_text(row).count("N/D*") == 2

    workbook = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )
    sheet = workbook["Modelo 4966"]
    labels = {
        sheet.cell(row=row_number, column=1).value: row_number
        for row_number in range(1, sheet.max_row + 1)
    }
    row_number = labels["Vencidos acima de 90 dias (conceito de arrasto)"]
    primary = sheet.cell(row=row_number, column=4)
    secondary = sheet.cell(row=row_number, column=5)
    assert primary.value == "N/D*"
    assert secondary.value == "N/D*"
    assert primary.comment is not None
    assert secondary.comment is not None
    assert primary.fill.patternType == "solid"
    assert secondary.fill.patternType == "solid"


def test_audit_dataframe_scales_only_currency_values_to_millions():
    audit = model_to_audit_dataframe(_model()).set_index("Indicador")

    assert audit.loc["Carteira total", "Dez/25 (R$ mm)"] == pytest.approx(270)
    assert audit.loc["Carteira total", "Dez/25 (%)"] == pytest.approx(100)
    assert audit.loc["PDD (Perda Esperada)", "Dez/25 (R$ mm)"] == pytest.approx(40)
    assert audit.loc["PDD / C5 (%)", "Dez/25 (%)"] == pytest.approx(40)
    assert format_brl_millions(1_234_567_890) == "1.235"


def test_percentage_format_preserves_rates_below_half_percent():
    assert format_percentage(0.0049, 1) == "0,5%"
    assert format_percentage(0.0005, 1) == "0,1%"
    assert format_percentage(-0.0001, 0) == "0%"


def test_excel_matches_row_spec_order_units_without_hatched_marker_column():
    workbook = openpyxl.load_workbook(BytesIO(build_carteira_4966_excel(_model())), data_only=True)

    assert workbook.sheetnames == ["Modelo 4966", "Alertas qualidade", "Glossário"]
    sheet = workbook["Modelo 4966"]
    assert sheet["A1"].value == TITLE
    assert sheet["A2"].value == "Indicador"
    assert sheet["A5"].value == "Em R$mm"
    assert all(
        cell.fill.patternType != "darkDown"
        for row in sheet.iter_rows()
        for cell in row
    )

    labels = [sheet.cell(row=row, column=1).value for row in range(1, sheet.max_row + 1)]
    rendered_rows = [label for label in labels if label in {spec.label for spec in ROW_SPECS}]
    assert rendered_rows == [spec.label for spec in ROW_SPECS]

    carteira_row = labels.index("Carteira total") + 1
    provision_row = labels.index("PDD (Perda Esperada)") + 1
    provision_ratio_row = labels.index("PDD / Carteira Total (%)") + 1
    assert sheet.cell(carteira_row, 6).value == pytest.approx(270)
    assert sheet.cell(carteira_row, 7).value == pytest.approx(1.0)
    assert sheet.cell(carteira_row, 7).number_format == "0.00%"
    assert sheet.cell(provision_row, 6).value == pytest.approx(40)
    assert sheet.cell(provision_row, 6).number_format == "#,##0"
    assert sheet.cell(provision_ratio_row, 6).value == pytest.approx(40 / 270)
    assert sheet.cell(provision_ratio_row, 6).number_format == "0.00%"
    assert sheet.cell(provision_ratio_row, 1).border.left.style is not None
    assert sheet.cell(provision_ratio_row, 2).border.left.style is not None
    assert sheet.cell(provision_ratio_row, 3).border.right.style is not None


def _model_with_pdd_ratio(pdd: float, total: float = 100.0):
    carteira = _carteira_frame().query("`Período` == '4/2025'").copy()
    carteira.loc[:, "Total Geral"] = total * MM
    ativo = _ativo_frame().query("`Período` == '4/2025'").copy()
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[0]] = -pdd * MM
    for column in EXPECTED_LOSS_COLUMNS[1:]:
        ativo.loc[:, column] = 0.0
    return build_carteira_4966_model(carteira, ativo, ["4/2025"])


def test_pdd_over_total_is_visible_and_uses_the_calculated_value():
    model = _model_with_pdd_ratio(40.0)
    rendered = render_carteira_4966_html(model)

    assert "PDD / Carteira Total (%)" in rendered
    assert "40,00%" in _visible_text(
        _html_row(rendered, "PDD / Carteira Total (%)")
    )

    workbook = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )
    cell = _excel_value_cell(workbook["Modelo 4966"], "PDD / Carteira Total (%)")
    assert cell.value == pytest.approx(0.4)
    assert cell.number_format == "0.00%"


@pytest.mark.parametrize("ativo_case", ["absent", "incomplete"])
def test_missing_or_incomplete_expected_loss_is_nd_star_with_diagnostic(ativo_case):
    carteira = _carteira_frame().query("`Período` == '4/2025'").copy()
    if ativo_case == "absent":
        ativo = pd.DataFrame()
    else:
        ativo = _ativo_frame().query("`Período` == '4/2025'").copy()
        ativo.loc[:, EXPECTED_LOSS_COLUMNS[-1]] = float("nan")

    model = build_carteira_4966_model(carteira, ativo, ["4/2025"])
    rendered = render_carteira_4966_html(model)

    assert model.cells["provision"]["4/2025"].primary is None
    assert model.cells["provision_over_portfolio"]["4/2025"].primary is None
    assert "N/D*" in _visible_text(_html_row(rendered, "PDD (Perda Esperada)"))
    ratio_row = _html_row(rendered, "PDD / Carteira Total (%)")
    assert "N/D*" in _visible_text(ratio_row)

    issue = model.pdd_quality_issue("4/2025")
    assert issue is not None
    assert issue.severity == "warning"
    diagnostic = quality_issue_message(issue)
    assert diagnostic.strip()
    assert any(
        token in diagnostic.lower()
        for token in ("relatório 2", "perda esperada", "incomplet", "indisponível")
    )
    assert html.escape(diagnostic, quote=True) in ratio_row

    workbook = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )
    sheet = workbook["Modelo 4966"]
    pdd_cell = _excel_value_cell(sheet, "PDD (Perda Esperada)")
    ratio_cell = _excel_value_cell(sheet, "PDD / Carteira Total (%)")
    assert pdd_cell.value == "N/D*"
    assert ratio_cell.value == "N/D*"
    assert pdd_cell.comment is not None
    assert ratio_cell.comment is not None
    assert "Alertas qualidade" in workbook.sheetnames
    assert workbook["Alertas qualidade"]["G2"].value


def test_pdd_ratio_above_attention_threshold_has_star_and_numeric_excel_value():
    model = _model_with_pdd_ratio(60.0)
    rendered = render_carteira_4966_html(model)
    ratio_row = _html_row(rendered, "PDD / Carteira Total (%)")

    assert "60,00%*" in _visible_text(ratio_row)
    assert "Alerta de confiabilidade" in ratio_row

    workbook = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )
    ratio_cell = _excel_value_cell(
        workbook["Modelo 4966"],
        "PDD / Carteira Total (%)",
    )
    assert ratio_cell.value == pytest.approx(0.6)
    assert ratio_cell.data_type == "n"
    assert "*" in ratio_cell.number_format
    assert ratio_cell.comment is not None

    alerts = workbook["Alertas qualidade"]
    assert alerts["B2"].value == "Atenção"
    assert alerts["F2"].value == pytest.approx(0.6)
    assert alerts["G2"].value

    raw_workbook = openpyxl.load_workbook(
        BytesIO(
            build_carteira_4966_raw_excel(
                model,
                _carteira_frame().query("`Período` == '4/2025'"),
                _ativo_frame().query("`Período` == '4/2025'"),
            )
        ),
        data_only=True,
    )
    raw_sheet = raw_workbook["Modelo calculado"]
    raw_headers = {
        raw_sheet.cell(row=1, column=column).value: column
        for column in range(1, raw_sheet.max_column + 1)
    }
    raw_labels = {
        raw_sheet.cell(row=row, column=1).value: row
        for row in range(2, raw_sheet.max_row + 1)
    }
    raw_cell = raw_sheet.cell(
        row=raw_labels["PDD / Carteira Total (%)"],
        column=raw_headers["Dez/25 (%)"],
    )
    assert raw_cell.value == pytest.approx(0.6)
    assert raw_cell.data_type == "n"
    assert "*" in raw_cell.number_format
    assert raw_cell.comment is not None
    assert "Alertas qualidade" in raw_workbook.sheetnames


@pytest.mark.parametrize(
    ("denominator_column", "row_key", "row_label"),
    [
        ("Total Geral", "provision_over_portfolio", "PDD / Carteira Total (%)"),
        ("C5", "provision_over_c5", "PDD / C5 (%)"),
        (
            "Inadimplência",
            "provision_over_delinquency",
            "PDD / Créditos vencidos acima de 90 dias (%)",
        ),
    ],
)
@pytest.mark.parametrize("denominator", [0.0, float("nan")], ids=["zero", "missing"])
def test_unavailable_ratio_denominator_is_nd_star_with_warning(
    denominator_column,
    row_key,
    row_label,
    denominator,
):
    carteira = _carteira_frame().query("`Período` == '4/2025'").copy()
    carteira.loc[:, denominator_column] = denominator
    ativo = _ativo_frame().query("`Período` == '4/2025'").copy()
    model = build_carteira_4966_model(carteira, ativo, ["4/2025"])

    assert model.cells[row_key]["4/2025"].primary is None
    row_html = _html_row(render_carteira_4966_html(model), row_label)
    assert "N/D*" in _visible_text(row_html)
    assert any(
        token in html.unescape(row_html).lower()
        for token in ("denominador", "zerad", "ausente", "indisponível")
    )

    workbook = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )
    ratio_cell = _excel_value_cell(workbook["Modelo 4966"], row_label)
    assert ratio_cell.value == "N/D*"
    assert ratio_cell.comment is not None
    assert any(
        token in ratio_cell.comment.text.lower()
        for token in ("denominador", "zerad", "ausente", "indisponível")
    )


def test_reliable_pdd_ratio_has_no_uncertainty_marker():
    model = _model_with_pdd_ratio(40.0)
    rendered = render_carteira_4966_html(model)
    ratio_row = _html_row(rendered, "PDD / Carteira Total (%)")

    assert _visible_text(ratio_row) == "40,00%"
    assert "Alerta de confiabilidade" not in ratio_row

    workbook = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )
    ratio_cell = _excel_value_cell(
        workbook["Modelo 4966"],
        "PDD / Carteira Total (%)",
    )
    assert ratio_cell.value == pytest.approx(0.4)
    assert ratio_cell.number_format == "0.00%"
    assert ratio_cell.comment is None
    assert "Alertas qualidade" not in workbook.sheetnames


def test_mixed_expected_loss_signs_keep_value_but_propagate_warning():
    carteira = _carteira_frame().query("`Período` == '4/2025'").copy()
    ativo = _ativo_frame().query("`Período` == '4/2025'").copy()
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[0]] = 1 * MM
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[1]] = -20 * MM
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[2]] = -3 * MM
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[3]] = -5 * MM

    model = build_carteira_4966_model(carteira, ativo, ["4/2025"])

    assert model.cells["provision"]["4/2025"].primary == pytest.approx(27 * MM)
    issue = model.cell_quality_issue("provision", "4/2025")
    assert issue is not None
    assert issue.code == "pdd_unexpected_sign"
    assert model.cell_quality_issue("provision_over_portfolio", "4/2025") == issue
    rendered = render_carteira_4966_html(model)
    assert "27*" in _visible_text(_html_row(rendered, "PDD (Perda Esperada)"))
    assert "27,00%*" in _visible_text(
        _html_row(rendered, "PDD / C5 (%)")
    )


def test_conflicting_expected_loss_rows_are_visible_but_flagged():
    carteira = _carteira_frame().query("`Período` == '4/2025'").copy()
    first = _ativo_frame().query("`Período` == '4/2025'").copy()
    second = first.copy()
    for column in EXPECTED_LOSS_COLUMNS[1:]:
        first.loc[:, column] = 0.0
        second.loc[:, column] = 0.0
    first.loc[:, EXPECTED_LOSS_COLUMNS[0]] = -10 * MM
    second.loc[:, EXPECTED_LOSS_COLUMNS[0]] = -40 * MM
    ativo = pd.concat([first, second], ignore_index=True)

    model = build_carteira_4966_model(carteira, ativo, ["4/2025"])

    assert model.cells["provision"]["4/2025"].primary == pytest.approx(40 * MM)
    issue = model.cell_quality_issue("provision", "4/2025")
    assert issue is not None
    assert issue.code == "pdd_conflicting_rows"
    pdd_row = _html_row(render_carteira_4966_html(model), "PDD (Perda Esperada)")
    assert "40*" in _visible_text(pdd_row)
    assert "valores conflitantes" in html.unescape(pdd_row)


def test_multiple_cell_issues_share_the_same_diagnostics_in_all_outputs():
    carteira = _carteira_frame().query("`Período` == '4/2025'").copy()
    carteira.loc[:, "Inadimplência"] = float("nan")
    ativo = _ativo_frame().query("`Período` == '4/2025'").copy()
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[0]] = 1 * MM
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[1]] = -20 * MM
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[2]] = -3 * MM
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[3]] = -5 * MM
    model = build_carteira_4966_model(carteira, ativo, ["4/2025"])

    issues = model.cell_quality_issues("provision_over_delinquency", "4/2025")
    assert {issue.code for issue in issues} == {
        "pdd_unexpected_sign",
        "missing_denominator",
    }
    rendered = render_carteira_4966_html(model)
    row_html = _html_row(
        rendered,
        "PDD / Créditos vencidos acima de 90 dias (%)",
    )
    assert 'aria-describedby="tc-4966-quality-provision_over_delinquency-4-2025"' in row_html
    assert "sinal positivo" in html.unescape(row_html)
    assert "denominador" in html.unescape(row_html)

    visual = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )["Modelo 4966"]
    visual_comment = _excel_value_cell(
        visual,
        "PDD / Créditos vencidos acima de 90 dias (%)",
    ).comment.text

    raw = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_raw_excel(model, carteira, ativo)),
        data_only=True,
    )["Modelo calculado"]
    raw_headers = {
        raw.cell(row=1, column=column).value: column
        for column in range(1, raw.max_column + 1)
    }
    raw_labels = {
        raw.cell(row=row, column=1).value: row
        for row in range(2, raw.max_row + 1)
    }
    raw_comment = raw.cell(
        row=raw_labels["PDD / Créditos vencidos acima de 90 dias (%)"],
        column=raw_headers["Dez/25 (%)"],
    ).comment.text
    assert visual_comment == raw_comment
    assert "sinal positivo" in raw_comment
    assert "denominador" in raw_comment

    source_alert = quality_issues_dataframe(model).query(
        "Checagem == 'pdd_unexpected_sign'"
    ).iloc[0]
    assert source_alert["Resultado (%)"] == pytest.approx(10.0)
    assert source_alert["PDD / Carteira Total (%)"] == pytest.approx(10.0)


def test_visual_and_raw_excel_keep_native_percentage_scale_and_two_decimals():
    carteira = _carteira_frame().query("`Período` == '4/2025'").copy()
    carteira.loc[:, "Total Geral"] = 100 * MM
    carteira.loc[:, "C1"] = 0.49 * MM
    ativo = _ativo_frame().query("`Período` == '4/2025'").copy()
    ativo.loc[:, EXPECTED_LOSS_COLUMNS[0]] = -120 * MM
    for column in EXPECTED_LOSS_COLUMNS[1:]:
        ativo.loc[:, column] = 0.0
    model = build_carteira_4966_model(carteira, ativo, ["4/2025"])

    rendered = render_carteira_4966_html(model)
    assert "0,49%" in rendered
    assert "120,00%" in rendered

    visual = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )["Modelo 4966"]
    visual_labels = {
        visual.cell(row=row, column=1).value: row
        for row in range(1, visual.max_row + 1)
    }
    c1_cell = visual.cell(row=visual_labels["C1"], column=3)
    pdd_ratio_cell = visual.cell(
        row=visual_labels["PDD / Carteira Total (%)"],
        column=2,
    )
    assert c1_cell.value == pytest.approx(0.0049)
    assert c1_cell.number_format == "0.00%"
    assert pdd_ratio_cell.value == pytest.approx(1.2)
    assert pdd_ratio_cell.number_format == '0.00%"*"'

    raw = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_raw_excel(model, carteira, ativo)),
        data_only=True,
    )
    raw_sheet = raw["Modelo calculado"]
    raw_headers = {
        raw_sheet.cell(row=1, column=column).value: column
        for column in range(1, raw_sheet.max_column + 1)
    }
    raw_labels = {
        raw_sheet.cell(row=row, column=1).value: row
        for row in range(2, raw_sheet.max_row + 1)
    }
    percent_column = raw_headers["Dez/25 (%)"]
    raw_c1 = raw_sheet.cell(row=raw_labels["C1"], column=percent_column)
    raw_pdd_ratio = raw_sheet.cell(
        row=raw_labels["PDD / Carteira Total (%)"],
        column=percent_column,
    )
    assert raw_c1.value == pytest.approx(0.0049)
    assert raw_c1.number_format == "0.00%"
    assert raw_pdd_ratio.value == pytest.approx(1.2)
    assert raw_pdd_ratio.number_format == '0.00%"*"'
    assert {"Alertas qualidade", "Rel16 Carteira", "Rel2 Ativo"}.issubset(
        raw.sheetnames
    )
    assert raw["Rel16 Carteira"]["B2"].value == "4/2025"
    assert raw["Rel2 Ativo"]["B2"].value == "4/2025"

    raw_without_ativo = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_raw_excel(model, carteira, pd.DataFrame())),
        data_only=True,
    )
    assert raw_without_ativo["Rel2 Ativo"]["A2"].value.startswith(
        "Dados de provisão indisponíveis"
    )


@pytest.mark.parametrize(
    ("pdd", "expected_severity"),
    [(40.0, None), (60.0, "warning"), (120.0, "critical")],
)
def test_pdd_quality_thresholds(pdd, expected_severity):
    model = _model_with_pdd_ratio(pdd)

    if expected_severity is None:
        assert model.quality_issues == ()
        assert quality_issues_dataframe(model).empty
        return

    issue = model.quality_issues[0]
    assert issue.severity == expected_severity
    assert issue.ratio == pytest.approx(pdd / 100)
    assert "Dez/25" in quality_issue_message(issue)


def test_pdd_above_portfolio_is_flagged_in_html_audit_and_excel():
    model = _model_with_pdd_ratio(120.0)
    rendered = render_carteira_4966_html(model)
    quality = quality_issues_dataframe(model)

    assert "tc-4966-quality-critical" in rendered
    assert "Alerta de confiabilidade" in rendered
    assert "120,00%" in rendered
    assert quality.loc[0, "Severidade"] == "Não confiável"
    assert quality.loc[0, "PDD / Carteira Total (%)"] == pytest.approx(120.0)

    workbook = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )
    assert workbook.sheetnames == ["Modelo 4966", "Alertas qualidade", "Glossário"]
    alerts = workbook["Alertas qualidade"]
    assert alerts["B2"].value == "Não confiável"
    assert alerts["F2"].value == pytest.approx(1.2)
    assert alerts["F2"].number_format == "0.00%"


def test_positive_pdd_with_zero_portfolio_is_unreliable_and_ratio_stays_missing():
    model = _model_with_pdd_ratio(10.0, total=0.0)

    issue = model.quality_issues[0]
    assert issue.code == "pdd_with_zero_portfolio"
    assert issue.severity == "critical"
    assert model.cells["provision_over_portfolio"]["4/2025"].primary is None
    assert "Carteira Total está zerada" in quality_issue_message(issue)


def test_negative_portfolio_is_unreliable_and_ratio_stays_missing():
    model = _model_with_pdd_ratio(10.0, total=-1.0)

    issue = model.quality_issues[0]
    assert issue.code == "negative_portfolio"
    assert issue.severity == "critical"
    assert model.cells["provision_over_portfolio"]["4/2025"].primary is None
    assert "Carteira Total é negativa" in quality_issue_message(issue)


def test_rounding_residue_above_portfolio_does_not_trigger_critical_alert():
    total = 73.61904855
    model = _model_with_pdd_ratio(total + 0.00000002, total=total)

    assert len(model.quality_issues) == 1
    assert model.quality_issues[0].severity == "warning"


def test_generic_expected_loss_is_ignored_and_all_four_leaf_values_are_required():
    carteira = _carteira_frame().query("`Período` == '4/2025'").copy()
    ativo = _ativo_frame().query("`Período` == '4/2025'").copy()
    ativo.loc[:, "Perda Esperada"] = -999 * MM

    complete = build_carteira_4966_model(carteira, ativo, ["4/2025"])
    assert complete.cells["provision"]["4/2025"].primary == pytest.approx(40 * MM)

    ativo.loc[:, EXPECTED_LOSS_COLUMNS[-1]] = float("nan")
    incomplete = build_carteira_4966_model(carteira, ativo, ["4/2025"])
    assert incomplete.cells["provision"]["4/2025"].primary is None
    assert incomplete.missing_provision_periods == ("4/2025",)


def test_small_portfolio_alert_keeps_useful_million_precision():
    model = _model_with_pdd_ratio(0.096, total=0.019)

    message = quality_issue_message(model.quality_issues[0])
    assert "R$ 0,096 mi" in message
    assert "R$ 0,019 mi" in message
