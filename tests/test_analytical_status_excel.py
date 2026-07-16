from io import BytesIO

import pytest
import xlsxwriter
from openpyxl import load_workbook

import app1


def _build_status_workbook(rows: list[dict]):
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    workbook.add_worksheet("base")
    app1._write_analytical_status_sheet(workbook, rows=rows)
    workbook.close()
    return load_workbook(BytesIO(output.getvalue()), data_only=True)


def test_analytical_status_writer_preserves_sheet_contract_and_number_formats():
    cases = [
        ("Perda Esperada / Est2+3", 0.1234, "0.0%"),
        ("Índice de Basileia", 0.1234, "0.00%"),
        ("percent", 0.1234, "0.0%"),
        ("percent_1", 0.1234, "0.0%"),
        ("percent_2", 0.1234, "0.00%"),
        ("Ativo/PL", 1.2345, "0.00x"),
        ("Crédito/PL (%)", 1.2345, "0.00x"),
        ("Carteira de Crédito Bruta / PL", 1.2345, "0.00x"),
        ("multiple", 1.2345, "0.00x"),
        ("money", 1234.56, "#,##0.00"),
        ("money", "1.234,56", "#,##0.00"),
        ("money", None, "General"),
    ]
    rows = [
        {
            "Instituição": f"Banco {index}",
            "Período": "Mar/26",
            "Indicador": format_key,
            "Valor": value,
            "Status analítico": "available",
            "Fonte analítica": "fonte",
            "Observação": "nota",
            "Format key": format_key,
        }
        for index, (format_key, value, _) in enumerate(cases, start=1)
    ]

    workbook = _build_status_workbook(rows)
    assert workbook.sheetnames == ["base", "status_analitico"]
    worksheet = workbook["status_analitico"]

    assert [worksheet.cell(row=1, column=column).value for column in range(1, 8)] == [
        "Instituição",
        "Período",
        "Indicador",
        "Valor",
        "Status analítico",
        "Fonte analítica",
        "Observação",
    ]
    assert worksheet.freeze_panes == "A2"
    assert [worksheet.column_dimensions[column].width for column in "ABCDEFG"] == pytest.approx(
        [28.7109375, 14.7109375, 32.7109375, 16.7109375, 24.7109375, 42.7109375, 72.7109375]
    )

    header = worksheet["A1"]
    assert header.font.bold is True
    assert header.font.color.rgb == "FFFFFFFF"
    assert header.fill.fgColor.rgb == "FF111111"
    assert (header.alignment.horizontal, header.alignment.vertical) == ("center", "center")
    assert header.border.left.style == "thin"
    assert header.border.left.color.rgb == "FFDDDDDD"

    text_cell = worksheet["A2"]
    assert (text_cell.alignment.horizontal, text_cell.alignment.vertical) == ("left", "top")
    assert text_cell.alignment.wrap_text is True
    assert text_cell.border.left.style == "thin"

    assert [worksheet.cell(row=row, column=1).value for row in range(2, 14)] == [
        f"Banco {index}" for index in range(1, 13)
    ]
    assert [worksheet.cell(row=row, column=3).value for row in range(2, 14)] == [
        format_key for format_key, _, _ in cases
    ]
    assert [worksheet.cell(row=row, column=4).value for row in range(2, 14)] == [
        0.1234,
        0.1234,
        0.1234,
        0.1234,
        0.1234,
        1.2345,
        1.2345,
        1.2345,
        1.2345,
        1234.56,
        1234.56,
        None,
    ]
    assert [worksheet.cell(row=row, column=4).number_format for row in range(2, 14)] == [
        number_format for _, _, number_format in cases
    ]


def test_analytical_status_writer_skips_empty_rows_without_creating_sheet():
    workbook = _build_status_workbook([])

    assert workbook.sheetnames == ["base"]
