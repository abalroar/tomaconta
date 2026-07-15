import html
import re
from io import BytesIO

import openpyxl
import pandas as pd
import pytest

from tabs.carteira_4966 import (
    EXPECTED_LOSS_COLUMNS,
    build_carteira_4966_excel,
    build_carteira_4966_model,
    build_carteira_4966_raw_excel,
    render_carteira_4966_html,
)


MM = 1_000_000.0
PERIOD = "4/2025"
PERIOD_LABEL = "Dez/25"
DELINQUENCY_LABEL = "Vencidos acima de 90 dias (conceito de arrasto)"


def _sources(delinquency: float | None) -> tuple[pd.DataFrame, pd.DataFrame]:
    carteira = pd.DataFrame(
        [
            {
                "Instituição": "BANCO TESTE - PRUDENCIAL",
                "Período": PERIOD,
                "C1": 10 * MM,
                "C2": 10 * MM,
                "C3": 10 * MM,
                "C4": 20 * MM,
                "C5": 40 * MM,
                "Total não Individualizado": 5 * MM,
                "Carteira não Informada ou não se Aplica": 5 * MM,
                "Total Exterior": 0.0,
                "Total Geral": 100 * MM,
                "Inadimplência": (
                    None if delinquency is None else delinquency * MM
                ),
            }
        ]
    )
    ativo = pd.DataFrame(
        [
            {
                "Instituição": "BANCO TESTE - PRUDENCIAL",
                "Período": PERIOD,
                EXPECTED_LOSS_COLUMNS[0]: -1 * MM,
                EXPECTED_LOSS_COLUMNS[1]: -1 * MM,
                EXPECTED_LOSS_COLUMNS[2]: -1 * MM,
                EXPECTED_LOSS_COLUMNS[3]: -1 * MM,
            }
        ]
    )
    return carteira, ativo


def _model(delinquency: float | None):
    carteira, ativo = _sources(delinquency)
    return build_carteira_4966_model(carteira, ativo, [PERIOD])


def _delinquency_html_cells(rendered: str) -> tuple[str, str]:
    row = re.search(
        r'<tr[^>]*data-row-key="delinquency"[^>]*>(.*?)</tr>',
        rendered,
        flags=re.DOTALL,
    )
    assert row is not None, "Linha de vencidos acima de 90 dias ausente no HTML"
    cells = re.findall(r"<td\b[^>]*>.*?</td>", row.group(1), flags=re.DOTALL)
    # O marcador estrutural pode ser removido junto com o hatch; os dois
    # últimos TDs são sempre valor e percentual do único período do fixture.
    assert len(cells) in (2, 3)
    return cells[-2], cells[-1]


def _visible_text(cell_html: str) -> str:
    without_screen_reader_text = re.sub(
        r'<span[^>]*class="[^"]*tc-4966-sr-only[^"]*"[^>]*>.*?</span>',
        "",
        cell_html,
        flags=re.DOTALL,
    )
    return re.sub(r"<[^>]+>", "", html.unescape(without_screen_reader_text)).strip()


def _row_for_label(sheet, *, label_column: int) -> int:
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=label_column).value == DELINQUENCY_LABEL:
            return row
    raise AssertionError(f"Linha {DELINQUENCY_LABEL!r} ausente no Excel")


def _visual_cells(workbook_bytes: bytes):
    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
    sheet = workbook["Modelo 4966"]
    row = _row_for_label(sheet, label_column=1)
    return workbook, sheet.cell(row=row, column=2), sheet.cell(row=row, column=3)


def _raw_cells(workbook_bytes: bytes):
    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
    sheet = workbook["Modelo calculado"]
    row = _row_for_label(sheet, label_column=1)
    headers = {
        sheet.cell(row=1, column=column).value: column
        for column in range(1, sheet.max_column + 1)
    }
    return (
        workbook,
        sheet.cell(row=row, column=headers[f"{PERIOD_LABEL} (R$ mm)"]),
        sheet.cell(row=row, column=headers[f"{PERIOD_LABEL} (%)"]),
    )


def _assert_alert_mentions_delinquency(workbook) -> None:
    assert "Alertas qualidade" in workbook.sheetnames
    sheet = workbook["Alertas qualidade"]
    values = " ".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ).lower()
    assert any(token in values for token in ("vencid", "inadimpl"))


def test_valid_delinquency_is_consistent_and_unmarked_in_all_exports():
    carteira, ativo = _sources(10.0)
    model = build_carteira_4966_model(carteira, ativo, [PERIOD])

    assert model.cells["delinquency"][PERIOD].primary == pytest.approx(10 * MM)
    assert model.cells["delinquency"][PERIOD].secondary == pytest.approx(0.10)
    assert model.cell_quality_issues("delinquency", PERIOD) == ()

    rendered = render_carteira_4966_html(model)
    amount_html, percent_html = _delinquency_html_cells(rendered)
    assert _visible_text(amount_html) == "10"
    assert _visible_text(percent_html) == "10,00%"
    assert "tc-4966-quality-" not in amount_html + percent_html
    assert "aria-describedby" not in amount_html + percent_html
    assert "*" not in _visible_text(amount_html + percent_html)
    assert "Inadimplência do Relatório 16" in rendered
    assert 'title="Carteira total do mesmo período"' in percent_html

    visual, visual_amount, visual_percent = _visual_cells(
        build_carteira_4966_excel(model)
    )
    assert visual_amount.value == pytest.approx(10.0)
    assert visual_amount.number_format == "#,##0"
    assert visual_percent.value == pytest.approx(0.10)
    assert visual_percent.number_format == "0.00%"
    assert visual_amount.fill.patternType is None
    assert visual_percent.fill.patternType is None
    assert visual_amount.comment is None
    percent_comment = visual_percent.comment.text if visual_percent.comment else ""
    percent_header = str(visual["Modelo 4966"]["C4"].value or "")
    assert (
        "mesmo período" in percent_comment
        or "carteira" in percent_header.lower()
    ), "O percentual deve explicitar que usa a carteira do mesmo período"
    assert "Alertas qualidade" not in visual.sheetnames

    raw, raw_amount, raw_percent = _raw_cells(
        build_carteira_4966_raw_excel(model, carteira, ativo)
    )
    assert raw_amount.value == pytest.approx(10.0)
    assert raw_amount.number_format == "#,##0.00"
    assert raw_percent.value == pytest.approx(0.10)
    assert raw_percent.number_format == "0.00%"
    assert raw_amount.comment is None
    assert raw_percent.comment is None
    assert "Alertas qualidade" not in raw.sheetnames


@pytest.mark.parametrize(
    ("delinquency", "expected_amount", "expected_ratio", "diagnostic_tokens"),
    [
        (None, None, None, ("indispon", "ausente", "publicad")),
        (-5.0, -5.0, -0.05, ("negativ",)),
        (120.0, 120.0, 1.20, ("super", "exced", "maior")),
    ],
    ids=["missing", "negative", "above_portfolio"],
)
def test_questionable_delinquency_marks_both_paired_cells_in_all_exports(
    delinquency,
    expected_amount,
    expected_ratio,
    diagnostic_tokens,
):
    carteira, ativo = _sources(delinquency)
    model = build_carteira_4966_model(carteira, ativo, [PERIOD])
    cell = model.cells["delinquency"][PERIOD]

    if expected_amount is None:
        assert cell.primary is None
        assert cell.secondary is None
    else:
        assert cell.primary == pytest.approx(expected_amount * MM)
        assert cell.secondary == pytest.approx(expected_ratio)

    issues = model.cell_quality_issues("delinquency", PERIOD)
    assert issues, "A própria linha de vencidos deve carregar o diagnóstico"
    diagnostic = " ".join(issue.code for issue in issues).lower() + " " + " ".join(
        str(issue.details) for issue in issues
    ).lower()

    rendered = render_carteira_4966_html(model)
    amount_html, percent_html = _delinquency_html_cells(rendered)
    for rendered_cell in (amount_html, percent_html):
        assert "tc-4966-quality-" in rendered_cell
        assert "*" in _visible_text(rendered_cell)
        assert re.search(r'aria-describedby="[^"]+"', rendered_cell), rendered_cell
        assert re.search(r'title="[^"]+"', rendered_cell), rendered_cell
        assert 'tabindex="0"' in rendered_cell
        assert 'role="tooltip"' in rendered_cell
    html_diagnostic = html.unescape(amount_html + percent_html).lower()
    assert any(token in html_diagnostic for token in diagnostic_tokens), diagnostic

    visual, visual_amount, visual_percent = _visual_cells(
        build_carteira_4966_excel(model)
    )
    raw, raw_amount, raw_percent = _raw_cells(
        build_carteira_4966_raw_excel(model, carteira, ativo)
    )

    for export_amount, export_percent in (
        (visual_amount, visual_percent),
        (raw_amount, raw_percent),
    ):
        if expected_amount is None:
            assert export_amount.value == "N/D*"
            assert export_percent.value == "N/D*"
        else:
            assert export_amount.value == pytest.approx(expected_amount)
            assert export_percent.value == pytest.approx(expected_ratio)
            assert "*" in export_amount.number_format
            assert "*" in export_percent.number_format
        assert export_amount.comment is not None
        assert export_percent.comment is not None
        assert export_amount.comment.text == export_percent.comment.text
        comment = export_amount.comment.text.lower()
        assert any(token in comment for token in diagnostic_tokens), comment

    assert visual_amount.fill.patternType == "solid"
    assert visual_percent.fill.patternType == "solid"
    _assert_alert_mentions_delinquency(visual)
    _assert_alert_mentions_delinquency(raw)


def test_html_and_visual_excel_do_not_use_structural_hatching():
    model = _model(10.0)
    rendered = render_carteira_4966_html(model)

    assert "repeating-linear-gradient" not in rendered

    workbook = openpyxl.load_workbook(
        BytesIO(build_carteira_4966_excel(model)),
        data_only=True,
    )
    sheet = workbook["Modelo 4966"]
    # Uma cor sólida pode ser usada, mas nenhuma célula estrutural deve ter
    # padrão rachurado/hachurado.
    assert all(
        sheet.cell(row=row, column=1).fill.patternType in (None, "solid")
        for row in range(1, sheet.max_row + 1)
    )
