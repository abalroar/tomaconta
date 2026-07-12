"""
Gera o tracker MVP do TomaConta em Excel — pronto para importar no Google Sheets.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT = "TomaConta_MVP_Tracker.xlsx"

# ── Cores ──────────────────────────────────────────────────────────────────
COLORS = {
    "Implementado":               ("C6EFCE", "375623"),
    "Pendente":                   ("FFC7CE", "9C0006"),
    "Parcialmente Implementado":  ("FFEB9C", "9C6500"),
    "Não classificável":          ("EDEDED", "636363"),
    "header_fill":                "1A1A1A",
    "header_font":                "FFFFFF",
    "aba_fill":                   "2C2C2C",
    "aba_font":                   "F5F5F5",
}

ORANGE = "F06611"   # Itaú BBA orange (accent lines)

# ── Dados ──────────────────────────────────────────────────────────────────
ROWS = [
    # (Aba, #, Pedido, Status, Observação)
    ("Rankings", "R1",  "Poder selecionar mais de um período",
     "Implementado", "Multiselect já disponível"),
    ("Rankings", "R2",  "Incluir ROE Trimestral como indicador",
     "Implementado", "ROE Trim. já consta na lista de 11 indicadores"),
    ("Rankings", "R3",  "Remover campo 'Ponderar média por'",
     "Implementado", "Campo não existe no Rankings (existe no Resumo/Snapshot) — já removido ou nunca estava ali"),
    ("Rankings", "R4",  "Manter linhas pontilhadas: 'média seleção' e 'média total SFN'",
     "Implementado", "Ambas as linhas tracejadas já existem com labels"),
    ("Rankings", "R5",  "Pools pré-selecionados: Top 5, Top 10, Top 20 (critério: total de ativos)",
     "Implementado", "Funcionando"),
    ("Rankings", "R6",  "Remover campo 'Ordenação', manter só direção (Maior→Menor / Menor→Maior)",
     "Implementado", "Campo 'Ordenação' não existe em Rankings — só o radio de direção"),
    ("Rankings", "R7",  "Ao selecionar período: opção 'tri vs tri anterior' E 'acumulado vs acumulado anterior'",
     "Implementado", "Ambas as sub-opções existem; sistema detecta consecutividade automaticamente"),
    ("Rankings", "R8",  "Cuidado com variação % para dados já em % (ROE, Basileia) — não calcular delta relativo sobre dado já percentual",
     "Pendente", "Não confirmado via análise estática — requer teste manual e possível fix na lógica de deltas"),
    ("Rankings", "R9–R12", "Seção (c) Visão Gráfica — conteúdo ilegível na transcrição",
     "Não classificável", "Imagens originais desfocadas neste trecho — realinhar com chefe antes da demo"),
    ("Rankings", "R_exp", "Não deixar opção de exportar na visão gráfica (Deltas / seção c)",
     "Pendente", "Exportação existe em todas as views atualmente; remover de views específicas requer ajuste de UI"),
    ("Rankings", "R13", "Tabela com pools pré-montados (Top 5 / 10 / 15 / 20)",
     "Implementado", "Pool selection funciona para a view de tabela"),
    ("Rankings", "R14", "Encaixar banco individual na posição correta dentro do ranking do pool",
     "Implementado", "Lógica existe: help text 'encaixá-las na posição real do ranking'"),
    ("Rankings", "R15", "Permitir exportar a tabela",
     "Implementado", "Download Excel disponível"),
    ("Rankings", "R16", "Formatação de números 'redondos' na tabela",
     "Pendente", "Não confirmado via análise — requer verificação visual e ajuste de formatação"),

    ("Taxa de Juros", "T1",  "Ranking aberto embaixo do gráfico com Top 10 e posições das casas selecionadas",
     "Implementado", "'📋 Ver ranking atual' existe, mostra Top 10 com posições — aprovado pelo chefe"),

    ("Contribuições FGC", "F1", "Retirar botões 'Top N' separados e incorporar como clusters (Top 5/10/15/20) no dropdown de Selecionar Instituições",
     "Pendente", "Atualmente é selectbox separado (10/20/50), não integrado como clusters ao seletor de instituições"),
    ("Contribuições FGC", "F2", "Escolher múltiplos períodos (mai/jun/set/dez) + toggle 'tri a tri' vs 'acumulados dos tris'",
     "Pendente", "Tem seleção de período inicial/final mas sem toggle tri vs acumulado"),
    ("Contribuições FGC", "F3", "Incluir LL Trimestral como variável adicional de escolha no FGC",
     "Pendente", "Não confirmado no tab de FGC — variável precisaria ser exposta na interface"),

    ("Scatter Plot", "S1", "LL Trimestral como opção de tamanho das bolinhas",
     "Implementado", "Todas as variáveis numéricas disponíveis para bubble size; LL Trim incluso"),
    ("Scatter Plot", "S2", "Bug: ao escolher 'variação de bastão', gráfico não renderiza",
     "Pendente", "Bug relatado pelo chefe; análise estática não encontrou causa — requer debug em runtime"),
    ("Scatter Plot", "S3", "Variáveis avançadas: PDD/Carteira, PDD/NI, Estágio 3/Carteira",
     "Pendente", "Chefe sinalizou como 'passo além' do MVP — manter no backlog para v2"),

    ("DRE", "D1", "3º dropdown 'base de comparação' (tri em questão vs acumulado)",
     "Parcialmente Implementado", "Existe dropdown 'Base lucro' (Acumulado vs Trimestral), mas não é o toggle tri-vs-acumulado completo pedido"),
    ("DRE", "D2", "Descer 'Desp PDD' e criar linha 'Resultado Intermediação Líquido Ajustado'",
     "Pendente", "Mudança de estrutura e ordem das variáveis — implica ajuste no layout da DRE"),
    ("DRE", "D3", "Abrir 'Desp PDD Outras Operações' como subconta recuada dentro de PDD",
     "Pendente", "Mudança de apresentação hierárquica na DRE"),
    ("DRE", "D4", "Verificar aderência de Desp PDD/NI e, se fizer sentido, levar ao Peer",
     "Pendente", "Condicional/analítico — depende de validação antes de implementar"),

    ("DRE Individual", "DI1", "Exibir nome do banco em vez do código da instituição",
     "Implementado", "DRE Individual já mostra nomes, não códigos"),

    ("Peer", "P1", "3º dropdown 'base de comparação' (tri em questão vs acumulado)",
     "Pendente", "Sem dropdown de base de comparação no Peer; apenas comparação lado a lado por período"),
    ("Peer", "P2", "Adicionar linhas 'Estágio 3 / Carteira' e 'Estágio 2 e 3 / Carteira'",
     "Pendente", "Existem valores absolutos e % individuais, mas não o ratio combinado Est2+3/Carteira"),
    ("Peer", "P3", "Adicionar 'Perda Esperada / Estágio 2 e 3'",
     "Pendente", "Só existe 'Perda Esperada / Estágio 3' atualmente"),
    ("Peer", "P4", "Incluir 'Índice Capital N1' (T1 = CET1 + AT1) na tabela",
     "Pendente", "Índice T1 existe no Rankings, mas não está na tabela do Peer"),

    ("Evolução", "E1", "Incluir Índice Capital N1 (T1) e ordenar: CET1 → N1 → BIS Total",
     "Implementado", "Tab Evolução já exibe CET1, T1 (N1) e Basileia Total nessa ordem"),

    ("Carteira 4966", "C1", "Em construção — menos prioritária",
     "Parcialmente Implementado", "Lógica base existe; dependente de dados do Relatório 16 — baixa prioridade confirmada pelo chefe"),
]

# ── Estatísticas ────────────────────────────────────────────────────────────
impl = sum(1 for r in ROWS if r[3] == "Implementado")
parcial = sum(1 for r in ROWS if r[3] == "Parcialmente Implementado")
pendente = sum(1 for r in ROWS if r[3] == "Pendente")
nclass = sum(1 for r in ROWS if r[3] == "Não classificável")
total_class = impl + parcial + pendente
completion = round((impl + 0.5 * parcial) / total_class * 100, 1)


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def make_font(hex_color, bold=False, size=11, name="Calibri"):
    return Font(color=hex_color, bold=bold, size=size, name=name)


def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MVP Tracker"

# ── Larguras ────────────────────────────────────────────────────────────────
col_widths = [22, 8, 62, 28, 72]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 14   # gap before title

# ── Banner de título ─────────────────────────────────────────────────────────
ws.merge_cells("A2:E2")
title_cell = ws["A2"]
title_cell.value = "TOMA CONTA — MVP TRACKER"
title_cell.fill = make_fill(COLORS["header_fill"])
title_cell.font = make_font(ORANGE, bold=True, size=16)
title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=2)
ws.row_dimensions[2].height = 32

# ── Linha de resumo ──────────────────────────────────────────────────────────
ws.merge_cells("A3:E3")
summary = ws["A3"]
summary.value = (
    f"Completion: {completion}%   |   "
    f"✅ Implementado: {impl}   |   "
    f"🟡 Parcial: {parcial}   |   "
    f"🔴 Pendente: {pendente}   |   "
    f"⚠️ Não classificável: {nclass}   |   "
    f"Demo: 15/04/2026"
)
summary.fill = make_fill("2C2C2C")
summary.font = make_font("F5F5F5", size=10)
summary.alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws.row_dimensions[3].height = 20

ws.row_dimensions[4].height = 6   # spacer

# ── Cabeçalhos ───────────────────────────────────────────────────────────────
HEADERS = ["Aba", "#", "Pedido", "Status", "Observação"]
for col, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=5, column=col, value=h)
    cell.fill = make_fill(COLORS["header_fill"])
    cell.font = make_font(COLORS["header_font"], bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border = thin_border()
ws.row_dimensions[5].height = 22

# ── Dados ────────────────────────────────────────────────────────────────────
current_aba = None
aba_start_row = None

for row_idx, (aba, num, pedido, status, obs) in enumerate(ROWS, start=6):
    fill_hex, font_hex = COLORS.get(status, ("FFFFFF", "000000"))

    for col_idx, val in enumerate([aba, num, pedido, status, obs], start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border()
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        if col_idx == 4:
            # Status column → colored
            cell.fill = PatternFill("solid", fgColor=fill_hex)
            cell.font = make_font(font_hex, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
        elif col_idx == 1:
            # Aba column
            if aba != current_aba:
                current_aba = aba
                cell.fill = make_fill(ORANGE)
                cell.font = make_font("FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           indent=1)
            else:
                cell.fill = make_fill("FAE8D8")
                cell.font = make_font("7A3700", size=10)
        elif col_idx == 2:
            cell.font = make_font("666666", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else:
            cell.font = Font(name="Calibri", size=10, color="1A1A1A")

    ws.row_dimensions[row_idx].height = 40

# ── Freeze & autofilter ──────────────────────────────────────────────────────
ws.freeze_panes = "A6"
ws.auto_filter.ref = f"A5:E{5 + len(ROWS)}"

# ── Aba de legenda ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Legenda")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 50

legend_data = [
    ("STATUS",             "SIGNIFICADO",                               "1A1A1A", "1A1A1A"),
    ("Implementado",       "Feature funcionando como descrito",         "C6EFCE", "375623"),
    ("Parcialmente Implementado", "Existe mas incompleto",              "FFEB9C", "9C6500"),
    ("Pendente",           "Não implementado — requer ação",            "FFC7CE", "9C0006"),
    ("Não classificável",  "Conteúdo ilegível na transcrição original", "EDEDED", "636363"),
]

for i, (label, desc, fill, fnt) in enumerate(legend_data, 1):
    c1 = ws2.cell(row=i, column=1, value=label)
    c2 = ws2.cell(row=i, column=2, value=desc)
    if i == 1:
        c1.fill = make_fill("1A1A1A")
        c1.font = make_font("FFFFFF", bold=True)
        c2.fill = make_fill("1A1A1A")
        c2.font = make_font("FFFFFF", bold=True)
    else:
        c1.fill = PatternFill("solid", fgColor=fill)
        c1.font = make_font(fnt, bold=True)
        c2.font = make_font("333333")
    for c in [c1, c2]:
        c.alignment = Alignment(vertical="center", indent=1)
        c.border = thin_border()
    ws2.row_dimensions[i].height = 22

wb.save(OUTPUT)
print(f"✅ Salvo: {OUTPUT}")
print(f"   {impl} Implementados | {parcial} Parciais | {pendente} Pendentes | {completion}% completion")
